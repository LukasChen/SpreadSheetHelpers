import sys
import os
import re
from openpyxl import load_workbook
from openpyxl.cell import MergedCell
from openpyxl.styles import stylesheet
from openpyxl.styles.numbers import BUILTIN_FORMATS, BUILTIN_FORMATS_MAX_SIZE

# MONKEY PATCH

def _expand_named_style(self, named_style):
    """
    Monkey-patched bind format definitions for a named style from the associated style
    record
    """
    try: xf = self.cellStyleXfs[named_style.xfId]
    except: return  # WORKAROUND for faulty Excel sheets
    named_style.font = self.fonts[xf.fontId]
    named_style.fill = self.fills[xf.fillId]
    named_style.border = self.borders[xf.borderId]
    if xf.numFmtId < BUILTIN_FORMATS_MAX_SIZE: formats = BUILTIN_FORMATS
    else:                                      formats = self.custom_formats
    if xf.numFmtId in formats:
        named_style.number_format = formats[xf.numFmtId]
    if xf.alignment:
        named_style.alignment = xf.alignment
    if xf.protection:
        named_style.protection = xf.protection

stylesheet.Stylesheet._expand_named_style = _expand_named_style  # monkey-patch


starting_row = 6
end_row = 16
start_col = 5
end_col = 40

template_filename = './template.xlsx'
output_filename = './output.xlsx'

if len(sys.argv) > 1:
    template_filename = sys.argv[1]
if len(sys.argv) > 2:
    output_filename = sys.argv[2]

print(template_filename, output_filename)

student_tasks = {}
weekdays = ['MON', 'TUE', 'WED', 'THU', 'FRI']

def read_student_tasks(student_name, full_filename):
    # Read data
    wb = load_workbook(full_filename)
    ws = wb.active
    student_tasks[student_name] = {}
    weekday = 'MON'
    for i in range(starting_row, end_row):
        weekdayHeader = ws.cell(row=i,column=start_col - 1).value;
        if weekdayHeader != None:
            weekday = weekdayHeader 


        for j in range(start_col, end_col):
            cell = ws.cell(row=i, column=j)
            if weekday not in student_tasks[student_name]:
                student_tasks[student_name][weekday] = []
            if not isinstance(cell, MergedCell):
                student_tasks[student_name][weekday].append(cell.value)


# Get src files
full_filenames = os.listdir(os.getcwd())

for file in full_filenames:
    match = re.search("\.xlsx$", file)

    if match:
        # print(file)
        filename = file.split('.')[0]
        if file != os.path.split(template_filename)[1] and file != os.path.split(output_filename)[1]:
            read_student_tasks(filename, file)



print(student_tasks)

output = load_workbook(template_filename, data_only=True)
output_sheet = output.active
student_search_row = 6
student_search_col = 12
student_end_row = 18
summary_interval = 26
data_col = 13
student_num = 6


for i in range(0, 5):
    base_row = i * summary_interval
    weekday = weekdays[i]
    print(weekday)
    
    name = ''
    index = 0
    for j in range(base_row + student_search_row, base_row + student_end_row):
        nameCell = output_sheet.cell(row=j, column=student_search_col).value
        if nameCell != None:
            index = 0
            if nameCell in student_tasks:
                name = nameCell
            else:
                continue
        for k in range(data_col, 48):
            cell = output_sheet.cell(row=j, column=k)
            if not isinstance(cell, MergedCell):
                cell.value = student_tasks[name][weekday][index]
                index += 1
        

output.save(output_filename)
