import sys
from colorama import Fore
import argparse
from dataclasses import dataclass
import openpyxl 

@dataclass
class Product:
    productPrice: int
    productQuantifier: str
    productQuantity: int
    productSum: int 

parser = argparse.ArgumentParser()

parser.add_argument("-i","--input", help = "输入文件")
parser.add_argument("-o","--output", help = "输出文件")

args = parser.parse_args()

if not args.output or not args.input:
    print(Fore.RED + "please specify input/output paths")
    sys.exit(1)


wbi = openpyxl.load_workbook(args.input)
wsi = wbi.active

massive_dict = {}

for i in range(1, wsi.max_row+1):
    cellId = wsi.cell(row = i, column=1)

    if type(cellId.value) is int:
        cellName = wsi.cell(row = i, column=2).value
        if cellName!= None:
            productPrice = wsi.cell(row = i, column=7).value
            productQuantifier = wsi.cell(row = i, column=3).value
            productQuantity = wsi.cell(row = i, column=4).value
            productSum = productPrice * productQuantity
            if cellName not in massive_dict:
                massive_dict[cellName] = Product(productPrice, productQuantifier,productQuantity, productSum)
            else:
                massive_dict[cellName].productQuantity += productQuantity
                massive_dict[cellName].productSum += productSum

print(massive_dict)

wbo = openpyxl.Workbook()
wso = wbo.active

wso["B1"] = "名称"
wso["C1"] = "品类"
wso["D1"] = "单位"
wso["E1"] = "数量"
wso["F1"] = "单价"
wso["G1"] = "金额"

currentRow = 2

for productName, product in massive_dict.items():
    wso.cell(row = currentRow, column = 2).value = productName
    wso.cell(row = currentRow, column = 4).value = product.productQuantifier
    wso.cell(row = currentRow, column = 5).value = product.productQuantity
    wso.cell(row = currentRow, column = 6).value = product.productPrice
    wso.cell(row = currentRow, column = 7).value = product.productSum
    currentRow += 1

wbo.save(args.output)

print('done')